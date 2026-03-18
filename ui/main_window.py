"""
MainWindow - ttkbootstrap版（i18n対応）
"""
import tkinter as tk
from tkinter import messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import calendar
import datetime
import os
import sys

from .task_dialog    import TaskDialog
from .options_dialog import OptionsDialog
from .date_picker    import DatePickerDialog
import i18n

def blend(base, accent, a=0.35):
    def h2r(h):
        h = h.lstrip("#")
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    br, bg, bb = h2r(base)
    ar, ag, ab = h2r(accent)
    return "#{:02x}{:02x}{:02x}".format(
        int(br+(ar-br)*a), int(bg+(ag-bg)*a), int(bb+(ab-bb)*a))

THEMES = {
    "dark":  {"BG":"#0D1117","SURFACE":"#161B22","CARD":"#1C2333",
              "BORDER":"#30363D","TEXT":"#E6EDF3","MUTED":"#7D8590"},
    "light": {"BG":"#F6F8FA","SURFACE":"#FFFFFF","CARD":"#EAEEF2",
              "BORDER":"#D0D7DE","TEXT":"#1F2328","MUTED":"#656D76"},
}


class MainWindow(tk.Frame):
    def __init__(self, master, store, config, app_ref):
        super().__init__(master, bg="#000000")
        self.store  = store
        self.config = config
        self.app    = app_ref

        today = datetime.date.today()
        self.year      = tk.IntVar(value=today.year)
        self.month     = tk.IntVar(value=today.month)
        self.day       = tk.IntVar(value=today.day)
        self.view_mode = tk.StringVar(value="month")  # 内部キー: month/day/year/custom
        self.custom_start = None
        self.custom_end   = None

        self._resolve_colors()
        self._build_menubar(master)
        self._build_header()
        self._build_body()
        self.refresh()

    def _t(self, key, default=""):
        return i18n.get(self.config, key, default)

    def _tl(self, key):
        return i18n.get_list(self.config, key)

    def _resolve_colors(self):
        mode = self.config.get("app_mode", "dark")
        t = THEMES.get(mode, THEMES["dark"])
        self.MODE    = mode
        self.BG      = t["BG"]
        self.SURFACE = t["SURFACE"]
        self.CARD    = t["CARD"]
        self.BORDER  = t["BORDER"]
        self.TEXT    = t["TEXT"]
        self.MUTED   = t["MUTED"]
        try:
            style = ttk.Style.instance
            self.ACCENT = style.colors.primary
        except Exception:
            self.ACCENT = "#4F7EFF"
        self.TODAY_BG = "#1F4A7A" if mode == "dark" else "#C8E0F8"
        self.TODAY_HL = "#4A90D9"
        self.DONE_CLR = "#495057" if mode == "dark" else "#adb5bd"
        try:
            self.configure(bg=self.BG)
        except Exception:
            pass

    def reload_theme(self):
        new_theme = "darkly" if self.config.get("app_mode","dark")=="dark" else "flatly"
        try:
            ttk.Style.instance.theme_use(new_theme)
        except Exception:
            pass
        self._resolve_colors()
        self._build_menubar(self.winfo_toplevel())
        self.hdr.destroy()
        self.body.destroy()
        self._build_header()
        self._build_body()
        # 言語切替後に表示ラベルを更新
        try:
            keys   = ["month","day","year","custom"]
            labels = [self._t("view_month"), self._t("view_day"),
                      self._t("view_year"),  self._t("view_custom")]
            cur_key = self.view_mode.get()
            if cur_key in keys:
                self._display_mode.set(labels[keys.index(cur_key)])
        except Exception:
            pass
        self.refresh()

    def _build_menubar(self, master):
        mb = tk.Menu(master)
        master.config(menu=mb)
        mb.add_command(label=self._t("menu_add_task"), command=self._open_add_task)
        opt = tk.Menu(mb, tearoff=0)
        opt.add_command(label=self._t("menu_settings"), command=self._open_options)
        mb.add_cascade(label=self._t("menu_options"), menu=opt)
        other = tk.Menu(mb, tearoff=0)
        other.add_command(label=self._t("menu_version"), command=self._show_version)
        other.add_separator()
        other.add_command(label=self._t("menu_export"), command=self._export_tasks)
        mb.add_cascade(label=self._t("menu_other"), menu=other)

    def _build_header(self):
        self.hdr = tk.Frame(self, bg=self.SURFACE, height=52)
        self.hdr.pack(fill="x")
        self.hdr.pack_propagate(False)

        left = tk.Frame(self.hdr, bg=self.SURFACE)
        left.pack(side="left", fill="y", padx=8, pady=6)

        tk.Button(left, text="◀", bg=self.SURFACE, fg=self.MUTED,
                  relief="flat", activebackground=self.BORDER,
                  activeforeground=self.TEXT, font=("Yu Gothic UI", 14),
                  bd=0, cursor="hand2", command=self._prev
                  ).pack(side="left", padx=2)

        self.title_label = tk.Label(left, text="", bg=self.SURFACE, fg=self.TEXT,
                                    font=("Yu Gothic UI", 14, "bold"), width=16, anchor="w")
        self.title_label.pack(side="left", padx=8)

        tk.Button(left, text="▶", bg=self.SURFACE, fg=self.MUTED,
                  relief="flat", activebackground=self.BORDER,
                  activeforeground=self.TEXT, font=("Yu Gothic UI", 14),
                  bd=0, cursor="hand2", command=self._next
                  ).pack(side="left", padx=2)

        tk.Button(left, text=self._t("btn_this_month"),
                  bg="#2563EB", fg="#FFFFFF",
                  activebackground="#1D4ED8", activeforeground="#FFFFFF",
                  font=("Yu Gothic UI", 10, "bold"),
                  relief="flat", cursor="hand2", padx=10, pady=4, bd=0,
                  command=self._goto_today).pack(side="left", padx=8)

        tk.Button(left, text=self._t("btn_today_tasks"),
                  bg=self.CARD, fg=self.TEXT,
                  activebackground=self.BORDER, activeforeground=self.TEXT,
                  font=("Yu Gothic UI", 10), relief="flat", cursor="hand2",
                  padx=10, pady=4, bd=0,
                  command=self._goto_day_view).pack(side="left", padx=4)

        right = tk.Frame(self.hdr, bg=self.SURFACE)
        right.pack(side="right", fill="y", padx=16, pady=6)

        tk.Label(right, text=self._t("lbl_view"), bg=self.SURFACE, fg=self.MUTED,
                 font=("Yu Gothic UI", 10)).pack(side="left", padx=(0, 4))

        style = ttk.Style()
        style.configure("White.TCombobox",
                         fieldbackground="white", background="white",
                         foreground="black", arrowcolor="black",
                         selectbackground="white", selectforeground="black",
                         padding=4)
        style.map("White.TCombobox",
                  fieldbackground=[("readonly","white")],
                  foreground=[("readonly","black")])

        # 内部キーと表示テキストを分離
        self._view_keys    = ["month","day","year","custom"]
        self._view_labels  = [self._t("view_month"), self._t("view_day"),
                              self._t("view_year"),  self._t("view_custom")]
        self._display_mode = tk.StringVar(
            value=self._view_labels[self._view_keys.index(self.view_mode.get())])
        self.view_combo = ttk.Combobox(
            right, textvariable=self._display_mode,
            values=self._view_labels, state="readonly", width=9,
            style="White.TCombobox", font=("Yu Gothic UI", 10)
        )
        self.view_combo.pack(side="left")
        self.view_combo.bind("<<ComboboxSelected>>", self._on_view_change)
        # 表示用変数の変化を内部キーに同期
        self._display_mode.trace_add("write", self._sync_view_mode)

    def _build_body(self):
        self.body = tk.Frame(self, bg=self.BG)
        self.body.pack(fill="both", expand=True)

    def _clear_body(self):
        for w in self.body.winfo_children():
            w.destroy()

    def _sync_view_mode(self, *_):
        """表示ラベルを内部キーに同期"""
        label = self._display_mode.get()
        if label in self._view_labels:
            self.view_mode.set(self._view_keys[self._view_labels.index(label)])

    def _on_view_change(self, *_):
        self._sync_view_mode()
        v = self.view_mode.get()
        if v == "custom":
            self._ask_custom_range()
        else:
            self.refresh()

    def _ask_custom_range(self):
        dlg = CustomRangeDialog(self.winfo_toplevel(),
                                self.custom_start, self.custom_end, self.config)
        self.winfo_toplevel().wait_window(dlg)
        if dlg.result_start and dlg.result_end:
            self.custom_start = dlg.result_start
            self.custom_end   = dlg.result_end
            self.refresh()
        else:
            self.view_mode.set("month")

    def refresh(self):
        self._resolve_colors()
        self._clear_body()
        v = self.view_mode.get()
        if v == "month":
            self._draw_month_view()
        elif v == "day":
            self._draw_day_view()
        elif v == "year":
            self._draw_year_view()
        elif v == "custom":
            if self.custom_start and self.custom_end:
                self._draw_custom_view()
            else:
                self._draw_month_view()

    # ══════════════════════════════════════════════════════
    #  月ビュー
    # ══════════════════════════════════════════════════════
    def _draw_month_view(self):
        y, m = self.year.get(), self.month.get()
        self.title_label.config(text=f"{y}年 {m}月" if self.config.get("language","ja")=="ja"
                                else f"{datetime.date(y,m,1).strftime('%B %Y')}", width=16)
        today     = datetime.date.today()
        first_dow = datetime.date(y, m, 1).weekday()
        days_in_m = calendar.monthrange(y, m)[1]
        tasks     = self.store.get_for_month(y, m)

        DAYS_JP = self._tl("days")
        C_TODAY_BG = "#1ABC9C"
        C_TODAY_FG = "#FFFFFF"
        C_DATE_BG  = "#2E86C1"
        C_DATE_FG  = "#FFFFFF"
        C_OTHER_BG = "#546E7A"
        C_OTHER_FG = "#B0BEC5"

        outer = tk.Frame(self.body, bg=self.BG)
        outer.pack(fill="both", expand=True, padx=10, pady=(4,10))

        SAT_FG = "#5B9BD5"
        SUN_FG = "#E05252"
        for col, day in enumerate(DAYS_JP):
            if col == 5:   hfg = SAT_FG
            elif col == 6: hfg = SUN_FG
            else:          hfg = "#CCCCCC" if self.MODE=="dark" else "#333333"
            tk.Label(outer, text=day, bg=self.SURFACE, fg=hfg,
                     font=("Yu Gothic UI", 10, "bold"), pady=5
                     ).grid(row=0, column=col, sticky="nsew", padx=1, pady=(0,2))
            outer.columnconfigure(col, weight=1)

        cells = {}
        for r in range(6):
            outer.rowconfigure(r+1, weight=1)
            for c in range(7):
                cell = tk.Frame(outer, bg=self.CARD,
                                highlightthickness=1,
                                highlightbackground=self.BORDER)
                cell.grid(row=r+1, column=c, sticky="nsew", padx=1, pady=1)
                cells[(r,c)] = cell

        prev_m    = m-1 if m > 1 else 12
        prev_y    = y   if m > 1 else y-1
        prev_days = calendar.monthrange(prev_y, prev_m)[1]

        date_labels = []
        slot = 0
        for r in range(6):
            for c in range(7):
                do = slot - first_dow
                if do < 0:
                    d  = datetime.date(prev_y, prev_m, prev_days+do+1)
                    nb, nf = C_OTHER_BG, C_OTHER_FG
                elif do >= days_in_m:
                    nm = m+1 if m < 12 else 1
                    ny = y   if m < 12 else y+1
                    d  = datetime.date(ny, nm, do-days_in_m+1)
                    nb, nf = C_OTHER_BG, C_OTHER_FG
                else:
                    d = datetime.date(y, m, do+1)
                    if d == today:
                        nb, nf = C_TODAY_BG, C_TODAY_FG
                    else:
                        nb, nf = C_DATE_BG, C_DATE_FG

                cell = cells[(r,c)]
                if d == today and d.month == m:
                    cell.config(highlightbackground="#1ABC9C", highlightthickness=2)

                num = tk.Label(cell, text=str(d.day), bg=nb, fg=nf,
                               font=("Yu Gothic UI", 9, "bold"),
                               anchor="ne", padx=4, pady=2)
                num.pack(fill="x")
                date_labels.append((num, nb, nf))
                num.bind("<Button-1>", lambda e, dt=d: self._open_add_task(dt))
                cell.bind("<Button-1>", lambda e, dt=d: self._open_add_task(dt))
                slot += 1

        def _force_colors():
            for lbl, bg, fg in date_labels:
                try: lbl.config(bg=bg, fg=fg)
                except Exception: pass
        outer.after_idle(_force_colors)
        outer.after(50, _force_colors)

        self._draw_task_bars(y, m, tasks, cells, first_dow, days_in_m)

    def _draw_task_bars(self, year, month, tasks, cells, first_dow, days_in_m):
        first_day = datetime.date(year, month, 1)
        last_day  = datetime.date(year, month, days_in_m)
        BAR_H     = 24
        BAR_PAD   = 2   # pady between bars

        def to_slot(d):
            return first_dow + (d - first_day).days

        # ── ステップ1：タスクごとに行番号を予約 ─────────────
        # cell_rows[(row, col)] = 次に使える行インデックス
        cell_rows = {}   # (row, col) -> list of used row_indices
        task_row  = {}   # task_id -> 行インデックス（複数日タスク固定用）

        def get_free_row(slots):
            """slotリスト全セルで空いている最小行を返す"""
            used = set()
            for sl in slots:
                r, c = sl // 7, sl % 7
                used |= set(cell_rows.get((r, c), []))
            row_idx = 0
            while row_idx in used:
                row_idx += 1
            return row_idx

        def reserve_row(slots, row_idx):
            for sl in slots:
                r, c = sl // 7, sl % 7
                if (r, c) not in cell_rows:
                    cell_rows[(r, c)] = []
                cell_rows[(r, c)].append(row_idx)

        # 複数日タスクを先に処理（行固定）
        multi_tasks  = []
        single_tasks = []
        for task in tasks:
            try:
                s = datetime.date.fromisoformat(task["start_date"])
                e = datetime.date.fromisoformat(task["due_date"])
            except Exception:
                continue
            if s != e:
                multi_tasks.append(task)
            else:
                single_tasks.append(task)

        # 複数日タスクはstart_dateでソート（行予約の安定性のため）
        multi_tasks.sort(key=lambda t: t.get("start_date",""))
        # 単日タスクはdue_date+timeでソート
        single_tasks.sort(key=lambda t: (t.get("due_date",""), t.get("due_time","23:59")))

        for task in multi_tasks + single_tasks:
            try:
                s = datetime.date.fromisoformat(task["start_date"])
                e = datetime.date.fromisoformat(task["due_date"])
            except Exception:
                continue
            sc = max(s, first_day)
            ec = min(e, last_day)

            # このタスクが占めるスロット一覧
            all_slots = []
            cur = sc
            while cur <= ec:
                sl = to_slot(cur)
                if 0 <= sl:
                    all_slots.append(sl)
                cur += datetime.timedelta(days=1)

            if not all_slots:
                continue

            row_idx = get_free_row(all_slots)
            task_row[task["id"]] = row_idx
            reserve_row(all_slots, row_idx)

        # ── ステップ2：行番号に基づいてバーを描画 ───────────
        bar_widgets = []

        for task in multi_tasks + single_tasks:
            try:
                s = datetime.date.fromisoformat(task["start_date"])
                e = datetime.date.fromisoformat(task["due_date"])
            except Exception:
                continue
            sc = max(s, first_day)
            ec = min(e, last_day)
            lc = task.get("label_color","")
            if task.get("done"):
                color = self.DONE_CLR
            elif lc and lc != "auto":
                color = str(lc)
            else:
                color = "#4F7EFF"
            due_time = task.get("due_time","")
            fg_text  = "#FFFFFF" if not task.get("done") else "#CCCCCC"
            row_idx  = task_row.get(task["id"], 0)
            # 行番号をpaddingに変換（日付ラベル分を考慮）
            bar_y    = row_idx * (BAR_H + BAR_PAD * 2) + 22  # 22px = 日付ラベル高さ

            cur = sc
            while cur <= ec:
                ss   = to_slot(cur)
                row  = ss // 7
                cs   = ss % 7
                send = min(ec, cur + datetime.timedelta(days=6-cs))
                ce   = to_slot(send) % 7
                is_start = (cur == sc)
                is_end   = (send == ec)

                for c in range(cs, ce+1):
                    if (row, c) not in cells:
                        continue
                    cell  = cells[(row, c)]
                    pad_l = 4 if c == cs else 0
                    pad_r = 4 if c == ce else 0

                    # place で絶対位置に配置（packではなく）
                    bar = tk.Frame(cell, bg=color, height=BAR_H, cursor="hand2")
                    bar.place(x=pad_l, y=bar_y,
                              relwidth=1.0,
                              width=-(pad_l + pad_r),
                              height=BAR_H)
                    bar_widgets.append((bar, color))

                    def _bind_all(w, t=task):
                        w.bind("<Button-1>", lambda e, tt=t: self._open_edit_task(tt))
                        w.bind("<Button-3>", lambda e, tt=t: self._task_context_menu(e, tt))

                    _bind_all(bar)

                    if c == cs:
                        if due_time and is_end and c == ce:
                            time_badge = tk.Label(bar,
                                                  text=f" {due_time} ",
                                                  bg=color, fg=fg_text,
                                                  font=("Yu Gothic UI", 8, "bold"),
                                                  anchor="w")
                            time_badge.pack(side="left", pady=2, padx=(0,2))
                            bar_widgets.append((time_badge, color))
                            _bind_all(time_badge)

                        title_lbl = tk.Label(bar, text=f" {task['title']}", bg=color,
                                             fg=fg_text, font=("Yu Gothic UI", 10, "bold"),
                                             anchor="w")
                        title_lbl.pack(side="left", fill="x", expand=True)
                        bar_widgets.append((title_lbl, color))
                        _bind_all(title_lbl)

                    elif c == ce and is_end and due_time:
                        time_badge = tk.Label(bar,
                                              text=f" {due_time} ",
                                              bg=color, fg=fg_text,
                                              font=("Yu Gothic UI", 8, "bold"),
                                              anchor="w")
                        time_badge.pack(side="left", pady=2)
                        bar_widgets.append((time_badge, color))
                        _bind_all(time_badge)
                    else:
                        _bind_all(bar)

                cur = send + datetime.timedelta(days=1)

        def _force_bar_colors():
            for w, col in bar_widgets:
                try: w.config(bg=col)
                except Exception: pass

        if bar_widgets:
            bar_widgets[0][0].after_idle(_force_bar_colors)
            bar_widgets[0][0].after(50, _force_bar_colors)

    # ══════════════════════════════════════════════════════
    #  日ビュー
    # ══════════════════════════════════════════════════════
    def _draw_day_view(self):
        y, m, d = self.year.get(), self.month.get(), self.day.get()
        today   = datetime.date(y, m, d)
        DAYS    = self._tl("days")
        if self.config.get("language","ja") == "ja":
            self.title_label.config(text=f"{y}年{m}月{d}日（{DAYS[today.weekday()]}）", width=16)
        else:
            self.title_label.config(text=today.strftime("%A, %B %d, %Y"), width=26)
        tasks = self.store.get_in_range(today, today)

        outer = tk.Frame(self.body, bg=self.BG)
        outer.pack(fill="both", expand=True, padx=20, pady=12)

        if not tasks:
            tk.Label(outer, text=self._t("no_tasks_today"),
                     bg=self.BG, fg=self.MUTED,
                     font=("Yu Gothic UI", 13)).pack(pady=40)
        else:
            for t in sorted(tasks, key=lambda t: (t.get("due_date",""), t.get("due_time","23:59"))):
                self._task_card_row(outer, t)

        ttk.Button(outer, text=self._t("btn_add_today"),
                   bootstyle="outline-primary",
                   command=lambda: self._open_add_task(today)
                   ).pack(pady=12)

    # ══════════════════════════════════════════════════════
    #  年ビュー
    # ══════════════════════════════════════════════════════
    def _draw_year_view(self):
        y = self.year.get()
        self.title_label.config(text=str(y), width=16)

        container = tk.Frame(self.body, bg=self.BG)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg=self.BG, highlightthickness=0)
        vsb    = ttk.Scrollbar(container, orient="vertical", command=canvas.yview, bootstyle="round")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner  = tk.Frame(canvas, bg=self.BG)
        win_id = canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        for idx in range(12):
            r, c = divmod(idx, 4)
            mini = self._mini_month(inner, y, idx+1)
            mini.grid(row=r, column=c, padx=8, pady=8, sticky="nsew")
            inner.columnconfigure(c, weight=1)

    def _mini_month(self, parent, year, month):
        today     = datetime.date.today()
        first_dow = datetime.date(year, month, 1).weekday()
        days_in_m = calendar.monthrange(year, month)[1]
        first_day = datetime.date(year, month, 1)
        last_day  = datetime.date(year, month, days_in_m)
        tasks     = self.store.get_for_month(year, month)
        DAYS      = self._tl("days")

        if self.config.get("language","ja") == "ja":
            month_label = f"{month}{self._t('month_suffix')}"
        else:
            month_label = datetime.date(year, month, 1).strftime("%B")

        outer = tk.Frame(parent, bg=self.CARD,
                         highlightthickness=1, highlightbackground=self.BORDER)
        tk.Label(outer, text=month_label, bg=self.ACCENT, fg="white",
                 font=("Yu Gothic UI", 11, "bold"), pady=5
                 ).grid(row=0, column=0, columnspan=7, sticky="ew")

        for ci in range(7):
            if ci == 5:   fg = "#5B9BD5"
            elif ci == 6: fg = "#E05252"
            else:         fg = "#CCCCCC" if self.MODE=="dark" else "#333333"
            tk.Label(outer, text=DAYS[ci], bg=self.SURFACE, fg=fg,
                     font=("Yu Gothic UI", 9, "bold"), pady=3
                     ).grid(row=1, column=ci, sticky="nsew", padx=1)
            outer.columnconfigure(ci, weight=1)

        cells = {}
        slot  = 0
        for ri in range(6):
            outer.rowconfigure(ri+2, weight=1)
            for ci in range(7):
                do = slot - first_dow
                if 0 <= do < days_in_m:
                    d  = datetime.date(year, month, do+1)
                    if ci == 5:   fg = "#5B9BD5"
                    elif ci == 6: fg = "#E05252"
                    else:         fg = self.TEXT
                    is_today = (d == today)
                    bg = self.TODAY_BG if is_today else self.CARD
                    hl = self.TODAY_HL if is_today else self.BORDER
                    cell = tk.Frame(outer, bg=bg,
                                    highlightthickness=1 if is_today else 0,
                                    highlightbackground=hl)
                    cell.grid(row=ri+2, column=ci, sticky="nsew", padx=1, pady=1)
                    cells[(ri,ci)] = (cell, d)
                    num = tk.Label(cell, text=str(do+1), bg=bg, fg=fg,
                                   font=("Yu Gothic UI", 9), anchor="ne", padx=3, pady=1)
                    num.pack(fill="x")
                    num.bind("<Button-1>", lambda e, yy=year, mm=month, dd=do+1:
                             self._jump_to_day(yy, mm, dd))
                    cell.bind("<Button-1>", lambda e, yy=year, mm=month, dd=do+1:
                              self._jump_to_day(yy, mm, dd))
                else:
                    cell = tk.Frame(outer, bg=self.BG)
                    cell.grid(row=ri+2, column=ci, sticky="nsew", padx=1, pady=1)
                    cells[(ri,ci)] = (cell, None)
                slot += 1

        def to_slot(d): return first_dow + (d - first_day).days
        for task in sorted(tasks, key=lambda t: t["start_date"]):
            try:
                s = datetime.date.fromisoformat(task["start_date"])
                e = datetime.date.fromisoformat(task["due_date"])
            except Exception:
                continue
            sc = max(s, first_day)
            ec = min(e, last_day)
            lc = task.get("label_color","")
            color = self.DONE_CLR if task.get("done") else (str(lc) if lc and lc!="auto" else "#4F7EFF")
            cur = sc
            while cur <= ec:
                ss   = to_slot(cur)
                row  = ss // 7
                cs   = ss % 7
                send = min(ec, cur + datetime.timedelta(days=6-cs))
                ce   = to_slot(send) % 7
                for c in range(cs, ce+1):
                    key = (row, c)
                    if key not in cells: continue
                    cell, cdate = cells[key]
                    if cdate is None: continue
                    bar = tk.Frame(cell, bg=color, height=5, cursor="hand2")
                    bar.pack(fill="x", padx=(2 if c==cs else 0, 2 if c==ce else 0))
                    bar.bind("<Button-1>", lambda e, t=task: self._open_edit_task(t))
                    bar.bind("<Button-3>", lambda e, t=task: self._task_context_menu(e, t))
                cur = send + datetime.timedelta(days=1)
        return outer

    def _jump_to_day(self, y, m, d):
        self.year.set(y); self.month.set(m); self.day.set(d)
        self.view_mode.set("day")
        self.refresh()

    # ══════════════════════════════════════════════════════
    #  カスタム期間ビュー
    # ══════════════════════════════════════════════════════
    def _draw_custom_view(self):
        s, e = self.custom_start, self.custom_end
        self.title_label.config(text=f"{s}  〜  {e}", width=16)
        tasks = self.store.get_in_range(s, e)
        delta = (e - s).days + 1
        outer = tk.Frame(self.body, bg=self.BG)
        outer.pack(fill="both", expand=True, padx=12, pady=8)

        if delta <= 31:
            DAYS = self._tl("days")
            canvas = tk.Canvas(outer, bg=self.BG, highlightthickness=0)
            vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview, bootstyle="round")
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y")
            canvas.pack(side="left", fill="both", expand=True)
            inner  = tk.Frame(canvas, bg=self.BG)
            wid    = canvas.create_window((0,0), window=inner, anchor="nw")
            inner.bind("<Configure>", lambda e2: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.bind("<Configure>", lambda e2: canvas.itemconfig(wid, width=e2.width))
            cur = s
            while cur <= e:
                day_tasks = [t for t in tasks
                             if datetime.date.fromisoformat(t["start_date"]) <= cur
                             <= datetime.date.fromisoformat(t["due_date"])]
                rf = tk.Frame(inner, bg=self.CARD,
                              highlightthickness=1, highlightbackground=self.BORDER)
                rf.pack(fill="x", pady=2)
                dow = cur.weekday()
                if dow == 5:   fg = "#5B9BD5"
                elif dow == 6: fg = "#E05252"
                else:          fg = self.TEXT
                label = f"{cur.month}/{cur.day}（{DAYS[dow]}）"
                tk.Label(rf, text=label, bg=self.CARD, fg=fg,
                         font=("Yu Gothic UI", 10, "bold"),
                         width=14, anchor="w", padx=8).pack(side="left")
                for t in day_tasks:
                    lc = t.get("label_color","")
                    color = self.DONE_CLR if t.get("done") else (str(lc) if lc and lc!="auto" else "#4F7EFF")
                    lbl = tk.Label(rf, text=f"  {t['title']}",
                                   bg=color, fg="white",
                                   font=("Yu Gothic UI", 9),
                                   padx=6, pady=2, cursor="hand2")
                    lbl.pack(side="left", padx=2)
                    lbl.bind("<Button-1>", lambda e2, tt=t: self._open_edit_task(tt))
                cur += datetime.timedelta(days=1)
        else:
            months = []
            cm = datetime.date(s.year, s.month, 1)
            em = datetime.date(e.year, e.month, 1)
            while cm <= em:
                months.append((cm.year, cm.month))
                ny = cm.year + (cm.month // 12)
                nm = cm.month % 12 + 1
                cm = datetime.date(ny, nm, 1)
            for idx, (my, mm) in enumerate(months):
                r, c = divmod(idx, 3)
                mini = self._mini_month(outer, my, mm)
                mini.grid(row=r, column=c, padx=8, pady=8, sticky="nsew")
                outer.columnconfigure(c, weight=1)

    # ══════════════════════════════════════════════════════
    #  共通パーツ
    # ══════════════════════════════════════════════════════
    def _task_color(self, task):
        if task.get("done"): return self.DONE_CLR
        lc = task.get("label_color","")
        if lc and lc != "auto": return lc
        return self.ACCENT

    def _task_card_row(self, parent, task):
        color    = self._task_color(task)
        is_done  = task.get("done", False)
        due_time = task.get("due_time", "")
        suffix   = self._t("time_suffix_day")

        card = tk.Frame(parent, bg=self.CARD,
                        highlightthickness=2, highlightbackground=color)
        card.pack(fill="x", pady=5)
        tk.Frame(card, bg=color, width=8).pack(side="left", fill="y")
        info = tk.Frame(card, bg=self.CARD)
        info.pack(side="left", fill="both", expand=True, padx=14, pady=10)

        title_text = ("✓  " if is_done else "") + task["title"]
        tk.Label(info, text=title_text, bg=self.CARD,
                 fg=self.MUTED if is_done else self.TEXT,
                 font=("Yu Gothic UI", 13, "bold"), anchor="w").pack(anchor="w")

        date_text = f"  {task['start_date']}  〜  {task['due_date']}"
        if due_time:
            date_text += f"   🕐 {due_time}{suffix}"
        tk.Label(info, text=date_text, bg=self.CARD,
                 fg=color if not is_done else self.MUTED,
                 font=("Yu Gothic UI", 11), anchor="w").pack(anchor="w", pady=(3,0))

        if task.get("detail"):
            detail = task["detail"]
            if len(detail) > 60: detail = detail[:60] + "…"
            tk.Label(info, text=f"  {detail}", bg=self.CARD,
                     fg=self.MUTED, font=("Yu Gothic UI", 10),
                     anchor="w").pack(anchor="w", pady=(2,0))

        badge_frame = tk.Frame(card, bg=self.CARD)
        badge_frame.pack(side="right", padx=12, pady=10)
        badge_color = "#43D9A2" if is_done else color
        badge_text  = self._t("badge_done") if is_done else self._t("badge_ongoing")
        tk.Label(badge_frame, text=badge_text,
                 bg=badge_color, fg="white",
                 font=("Yu Gothic UI", 9, "bold"),
                 padx=8, pady=3).pack()

        for w in (card, info, badge_frame):
            w.bind("<Button-1>", lambda e, t=task: self._open_edit_task(t))
            w.bind("<Button-3>", lambda e, t=task: self._task_context_menu(e, t))

    def _task_context_menu(self, event, task):
        menu = tk.Menu(self, tearoff=0)
        label = self._t("ctx_complete") if not task.get("done") else self._t("ctx_revert")
        menu.add_command(label=label, command=lambda: self._toggle_done(task))
        menu.add_command(label=self._t("ctx_edit"),   command=lambda: self._open_edit_task(task))
        menu.add_separator()
        menu.add_command(label=self._t("ctx_delete"), command=lambda: self._delete_task(task))
        menu.tk_popup(event.x_root, event.y_root)

    def _toggle_done(self, task):
        self.store.update(task["id"], done=not task.get("done", False))
        self.refresh()

    def _delete_task(self, task):
        if messagebox.askyesno(self._t("confirm_delete"),
                               f"「{task['title']}」{self._t('confirm_delete_msg')}",
                               parent=self.winfo_toplevel()):
            self.store.delete(task["id"])
            self.refresh()

    def _prev(self):
        v = self.view_mode.get()
        if v == "month":
            m, y = self.month.get(), self.year.get()
            if m == 1: self.month.set(12); self.year.set(y-1)
            else:      self.month.set(m-1)
        elif v == "day":
            d = datetime.date(self.year.get(), self.month.get(), self.day.get())
            d -= datetime.timedelta(days=1)
            self.year.set(d.year); self.month.set(d.month); self.day.set(d.day)
        elif v == "year":
            self.year.set(self.year.get()-1)
        self.refresh()

    def _next(self):
        v = self.view_mode.get()
        if v == "month":
            m, y = self.month.get(), self.year.get()
            if m == 12: self.month.set(1); self.year.set(y+1)
            else:        self.month.set(m+1)
        elif v == "day":
            d = datetime.date(self.year.get(), self.month.get(), self.day.get())
            d += datetime.timedelta(days=1)
            self.year.set(d.year); self.month.set(d.month); self.day.set(d.day)
        elif v == "year":
            self.year.set(self.year.get()+1)
        self.refresh()

    def _goto_today(self):
        t = datetime.date.today()
        self.year.set(t.year); self.month.set(t.month); self.day.set(t.day)
        self.view_mode.set("month")
        self.refresh()

    def _goto_day_view(self):
        t = datetime.date.today()
        self.year.set(t.year); self.month.set(t.month); self.day.set(t.day)
        self.view_mode.set("day")
        self.refresh()

    def _open_add_task(self, default_start=None):
        dlg = TaskDialog(self.winfo_toplevel(), self.store,
                         config=self.config, default_start=default_start)
        self.winfo_toplevel().wait_window(dlg)
        self.refresh()

    def _open_edit_task(self, task):
        dlg = TaskDialog(self.winfo_toplevel(), self.store,
                         config=self.config, task=task)
        self.winfo_toplevel().wait_window(dlg)
        self.refresh()

    def _open_options(self):
        dlg = OptionsDialog(self.winfo_toplevel(), self.config, self.app)
        self.winfo_toplevel().wait_window(dlg)
        self.reload_theme()

    def _show_version(self):
        from app import APP_VERSION, APP_NAME
        dlg = tk.Toplevel(self.winfo_toplevel())
        dlg.withdraw()
        dlg.title(self._t("ver_title"))
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self.winfo_toplevel())

        base = sys._MEIPASS if getattr(sys,"frozen",False) else \
               os.path.dirname(os.path.abspath(sys.modules["app"].__file__))
        ico = os.path.join(base, "assets", "icon.ico")
        if os.path.exists(ico):
            try: dlg.iconbitmap(ico)
            except Exception: pass

        outer = tk.Frame(dlg, bg=self.BG, padx=30, pady=24)
        outer.pack()

        png = os.path.join(base, "assets", "icon.png")
        if os.path.exists(png):
            try:
                from PIL import Image, ImageTk
                img   = Image.open(png).resize((48,48), Image.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                lbl   = tk.Label(outer, image=photo, bg=self.BG)
                lbl.image = photo
                lbl.pack(pady=(0,10))
            except Exception: pass

        tk.Label(outer, text=APP_NAME, bg=self.BG, fg=self.TEXT,
                 font=("Yu Gothic UI", 14, "bold")).pack()
        tk.Label(outer, text=self._t("ver_version")+APP_VERSION,
                 bg=self.BG, fg=self.MUTED, font=("Yu Gothic UI", 10)).pack(pady=(4,0))
        tk.Label(outer, text=self._t("ver_copyright"),
                 bg=self.BG, fg=self.MUTED, font=("Yu Gothic UI", 9)).pack(pady=(2,12))
        ttk.Button(outer, text=self._t("btn_ok"), bootstyle="primary",
                   width=10, command=dlg.destroy).pack()

        dlg.update_idletasks()
        pw = self.winfo_toplevel().winfo_rootx() + self.winfo_toplevel().winfo_width()//2
        ph = self.winfo_toplevel().winfo_rooty() + self.winfo_toplevel().winfo_height()//2
        w, h = dlg.winfo_width(), dlg.winfo_height()
        dlg.geometry(f"+{pw-w//2}+{ph-h//2}")
        dlg.deiconify()

    def _export_tasks(self):
        from collections import defaultdict
        tasks = self.store.get_all()
        if not tasks:
            messagebox.showinfo(self._t("export_done"), self._t("export_none"),
                                parent=self.winfo_toplevel())
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(self._t("export_err_title"), self._t("export_err_msg"),
                                 parent=self.winfo_toplevel())
            return

        by_month = defaultdict(list)
        for t in tasks:
            ym = t.get("start_date","")[:7]
            if ym: by_month[ym].append(t)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        HDR_FILL  = PatternFill("solid", fgColor="1565C0")
        DONE_FILL = PatternFill("solid", fgColor="EEEEEE")
        ALT_FILL  = PatternFill("solid", fgColor="F0F7FF")
        HDR_FONT  = Font(name="Yu Gothic UI", bold=True, color="FFFFFF", size=10)
        BODY_FONT = Font(name="Yu Gothic UI", size=10)
        DONE_FONT = Font(name="Yu Gothic UI", size=10, color="999999")
        CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin      = Side(style="thin", color="CCCCCC")
        BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

        HEADERS = self._tl("export_headers")
        COL_W   = [5, 28, 35, 12, 12, 9, 20, 10]

        for ym in sorted(by_month.keys()):
            y2, mm = ym.split("-")
            if self.config.get("language","ja") == "ja":
                sheet_name = f"{y2}年{int(mm)}月"
                title_text = self._t("export_sheet_title") + f"{y2}年{int(mm)}月"
            else:
                import calendar as cal_mod
                month_name = cal_mod.month_name[int(mm)]
                sheet_name = f"{month_name} {y2}"
                title_text = self._t("export_sheet_title") + f"{month_name} {y2}"

            ws = wb.create_sheet(title=sheet_name)
            ws.merge_cells("A1:H1")
            c = ws["A1"]
            c.value = title_text
            c.font  = Font(name="Yu Gothic UI", bold=True, size=13, color="1565C0")
            c.alignment = CENTER
            ws.row_dimensions[1].height = 28

            for ci, (hdr, w) in enumerate(zip(HEADERS, COL_W), start=1):
                cell = ws.cell(row=2, column=ci, value=hdr)
                cell.fill = HDR_FILL; cell.font = HDR_FONT
                cell.alignment = CENTER; cell.border = BORDER
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.row_dimensions[2].height = 20

            month_tasks = sorted(by_month[ym],
                                  key=lambda t: (t.get("due_date",""), t.get("due_time","23:59")))
            for ri, t in enumerate(month_tasks, start=3):
                is_done = t.get("done", False)
                fill    = DONE_FILL if is_done else (ALT_FILL if ri%2==0 else None)
                font    = DONE_FONT if is_done else BODY_FONT
                status  = self._t("export_done_status") if is_done else self._t("export_ongoing")
                row_data = [ri-2, t.get("title",""), t.get("detail",""),
                            t.get("start_date",""), t.get("due_date",""),
                            t.get("due_time",""), t.get("note",""), status]
                for ci, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = font; cell.border = BORDER
                    cell.alignment = CENTER if ci in (1,4,5,6,8) else LEFT
                    if fill: cell.fill = fill
                ws.row_dimensions[ri].height = 18
            ws.freeze_panes = "A3"

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        os.makedirs(desktop, exist_ok=True)
        now   = datetime.datetime.now()
        fname = f"TaskPlovo_{now.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        fpath = os.path.join(desktop, fname)
        wb.save(fpath)
        try: os.startfile(fpath)
        except Exception: pass
        self._show_info_dialog(self._t("export_done"),
                               self._t("export_saved") + fname)

    def _show_info_dialog(self, title, message):
        """アプリアイコン付きの情報ダイアログ"""
        dlg = tk.Toplevel(self.winfo_toplevel())
        dlg.withdraw()
        dlg.title(title)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self.winfo_toplevel())

        base = sys._MEIPASS if getattr(sys, "frozen", False) else \
               os.path.dirname(os.path.abspath(sys.modules["app"].__file__))
        ico = os.path.join(base, "assets", "icon.ico")
        if os.path.exists(ico):
            try: dlg.iconbitmap(ico)
            except Exception: pass

        outer = tk.Frame(dlg, bg=self.BG, padx=28, pady=20)
        outer.pack()
        tk.Label(outer, text=message, bg=self.BG, fg=self.TEXT,
                 font=("Yu Gothic UI", 10), justify="left").pack(pady=(0, 16))
        ttk.Button(outer, text=self._t("btn_ok"), bootstyle="primary",
                   width=10, command=dlg.destroy).pack()

        dlg.update_idletasks()
        pw = self.winfo_toplevel().winfo_rootx() + self.winfo_toplevel().winfo_width()  // 2
        ph = self.winfo_toplevel().winfo_rooty() + self.winfo_toplevel().winfo_height() // 2
        w, h = dlg.winfo_width(), dlg.winfo_height()
        dlg.geometry(f"+{pw-w//2}+{ph-h//2}")
        dlg.deiconify()
        dlg.wait_window()


# ══════════════════════════════════════════════════════════
#  カスタム期間ダイアログ
# ══════════════════════════════════════════════════════════
class CustomRangeDialog(ttk.Toplevel):
    def __init__(self, parent, init_start=None, init_end=None, config=None):
        super().__init__(parent)
        self.withdraw()
        self._cfg = config or {}
        self.result_start = None
        self.result_end   = None
        self._start = tk.StringVar(value=init_start.isoformat() if init_start else "")
        self._end   = tk.StringVar(value=init_end.isoformat()   if init_end   else "")
        self.title(i18n.get(self._cfg, "custom_title"))
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self.geometry("380x200")

        base = sys._MEIPASS if getattr(sys,"frozen",False) else \
               os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        ico = os.path.join(base, "assets", "icon.ico")
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        pw = parent.winfo_rootx() + parent.winfo_width()//2
        ph = parent.winfo_rooty() + parent.winfo_height()//2
        self.geometry(f"+{pw-self.winfo_width()//2}+{ph-self.winfo_height()//2}")
        self.deiconify()

    def _t(self, key): return i18n.get(self._cfg, key)

    def _build(self):
        ttk.Label(self, text=self._t("custom_prompt"),
                  font=("Yu Gothic UI",11,"bold")).pack(padx=24, pady=(20,14))
        row = ttk.Frame(self)
        row.pack(fill="x", padx=24)

        def date_row(parent, label, var):
            f = ttk.Frame(parent)
            f.pack(side="left", expand=True, fill="x", padx=(0,8))
            ttk.Label(f, text=label, font=("Yu Gothic UI",9)).pack(anchor="w")
            box = ttk.Frame(f)
            box.pack(fill="x")
            lbl = ttk.Label(box, textvariable=var, font=("Yu Gothic UI",10), width=12, anchor="w")
            lbl.pack(side="left", padx=8, pady=6)
            btn = ttk.Label(box, text="📅", cursor="hand2")
            btn.pack(side="right", padx=6)
            for w in (box, lbl, btn):
                w.bind("<Button-1>", lambda e, v=var: self._pick(v))

        date_row(row, self._t("custom_start"), self._start)
        date_row(row, self._t("custom_end"),   self._end)

        btn_row = ttk.Frame(self)
        btn_row.pack(fill="x", padx=24, pady=(16,0))
        ttk.Button(btn_row, text=self._t("btn_cancel"), bootstyle="outline-secondary",
                   command=self.destroy).pack(side="right", padx=(8,0))
        ttk.Button(btn_row, text=self._t("btn_apply"), bootstyle="primary",
                   command=self._apply).pack(side="right")

    def _pick(self, var):
        try: init = datetime.date.fromisoformat(var.get())
        except Exception: init = datetime.date.today()
        dlg = DatePickerDialog(self, initial_date=init)
        self.wait_window(dlg)
        if dlg.result: var.set(dlg.result.isoformat())

    def _apply(self):
        try:
            s = datetime.date.fromisoformat(self._start.get())
            e = datetime.date.fromisoformat(self._end.get())
            if e < s:
                messagebox.showwarning(self._t("err_input"),
                                       self._t("err_date_order"), parent=self)
                return
            self.result_start = s; self.result_end = e
            self.destroy()
        except Exception:
            messagebox.showwarning(self._t("err_input"),
                                   self._t("err_start_empty"), parent=self)
